import datetime

from django.core.mail import send_mass_mail
from django.core.mail import get_connection, EmailMultiAlternatives
from threading import Thread

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from openpyxl import Workbook
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference,
    BarChart
)
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Border, Side
from collections import Counter

from api.models import Answer


def send_mass_html_mail(datatuple, fail_silently=False, user=None, password=None, connection=None):
    """
    Given a datatuple of (subject, text_content, html_content, from_email,
    recipient_list), sends each message to each recipient list. Returns the
    number of emails sent.

    If from_email is None, the DEFAULT_FROM_EMAIL setting is used.
    If auth_user and auth_password are set, they're used to log in.
    If auth_user is None, the EMAIL_HOST_USER setting is used.
    If auth_password is None, the EMAIL_HOST_PASSWORD setting is used.

    """
    connection = connection or get_connection(
        username=user, password=password, fail_silently=fail_silently)
    messages = []
    for subject, text, html, from_email, recipient in datatuple:
        message = EmailMultiAlternatives(subject, text, from_email, recipient)
        message.attach_alternative(html, 'text/html')
        messages.append(message)
    return connection.send_messages(messages)


def send_my_mass_mail(survey_id, survey_title, email_list, html=True) -> None:
    """
    starts new thread sending mass email (to prevent API freeze) - significantly speeds up the request
    current link to survey: DOMAIN_NAME/survey/<survey_id>
    """
    # partial_link = reverse('surveys-uuid', args=[survey_id])
    # partial_link = partial_link.replace('surveys', 'survey')
    # ^ pointless

    partial_link = f'/survey/{survey_id}'
    survey_link = settings.DOMAIN_NAME + partial_link

    context = {'link': survey_link}
    html_message = render_to_string('email_template.html', context=context)
    txt_message = strip_tags(html_message)

    if html:
        data_tuple_html = ((survey_title, txt_message, html_message, None, email_list),)
        t = Thread(target=send_mass_html_mail, args=(data_tuple_html,))
        t.start()
    else:
        data_tuple_txt = ((survey_title, txt_message, None, email_list),)
        t = Thread(target=send_mass_mail, args=(data_tuple_txt,))
        t.start()


def xlsx_question_charts_file(queryset, question_val, answer_type) -> HttpResponse:
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={question}-{date}-results.xlsx'.format(
        date=datetime.datetime.now().strftime('%Y-%m-%d'), question=question_val
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f'"{question_val}" pie'

    if answer_type == 'option':
        answers = queryset.prefetch_related('option').values_list('option__content', flat=True)
    elif answer_type == 'content_numeric':
        answers = queryset.values_list('content_numeric', flat=True)
    else:
        answers = queryset.values_list('content_character', flat=True)
        answers = [' '.join(content_character.lower().strip().split()) for content_character in answers
                   if content_character is not None]

    counted = Counter(answers)
    statement = len(counted) > 5
    counted = counted.most_common(5) if statement else counted.most_common()
    max_row = len(counted)

    if not max_row:
        row = ['Brak odpowiedzi na to pytanie']
        worksheet.append(row)
        workbook.save(response)
        return response

    # normal pie
    for row in counted:
        worksheet.append(row)
    pie = PieChart()
    labels = Reference(worksheet, min_col=1, min_row=1, max_row=max_row)
    data = Reference(worksheet, min_col=2, min_row=1, max_row=max_row)
    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = question_val
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = True
    worksheet.add_chart(pie, "D1")

    # projected pie
    ws = workbook.create_sheet(title=f'"{question_val}" projection')
    for row in counted:
        ws.append(row)
    projected_pie = ProjectedPieChart()
    projected_pie.type = "bar"
    projected_pie.splitType = "pos"  # split by value
    labels = Reference(ws, min_col=1, min_row=1, max_row=max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
    projected_pie.add_data(data)
    projected_pie.dataLabels = DataLabelList()
    projected_pie.dataLabels.showVal = True
    projected_pie.set_categories(labels)
    projected_pie.title = question_val
    ws.add_chart(projected_pie, "D2")

    # bar
    ws2 = workbook.create_sheet(title=f'"{question_val}" bar')
    for row in counted:
        ws2.append(row)
    bar = BarChart()
    bar.type = "col"
    # bar.style = 10
    bar.title = question_val
    bar.y_axis.title = "Ilość wystąpień"
    bar.x_axis.title = "Udzielone odpowiedzi"
    labels = Reference(ws2, min_col=1, min_row=1, max_row=max_row)
    data = Reference(ws2, min_col=2, min_row=1, max_row=max_row)
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.add_data(data)
    bar.set_categories(labels)
    bar.shape = 4
    ws2.add_chart(bar, "D3")

    workbook.save(response)
    return response


def xlsx_survey_results(queryset, survey_title) -> HttpResponse:
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={survey_title}-{date}-results.xlsx'.format(
        date=datetime.datetime.now().strftime('%Y-%m-%d'), survey_title=survey_title
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f'"{survey_title}" results'

    for col_num, question in enumerate(queryset, 1):
        row_num = 1
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = question.value
        cell.font = Font(name='Calibri', bold=True)
        cell.border = Border(bottom=Side(border_style='medium', color='FF000000'), )
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 30

        for answer in Answer.objects.filter(question_id=question.id):
            row_num += 1
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = answer.get_content_type_value(question.get_answer_content_type())

    worksheet.freeze_panes = worksheet['A2']
    workbook.save(response)
    return response


# def xlsx_survey_results2(queryset, queryset_combined, survey_title) -> HttpResponse:    # co lepsze?
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename={survey_title}-{date}-results.xlsx'.format(
#         date=datetime.datetime.now().strftime('%Y-%m-%d'), survey_title=survey_title
#     )
#     workbook = Workbook()
#     worksheet = workbook.active
#     worksheet.title = f'"{survey_title}" results'
#
#     survey_data = {question: [] for question in queryset}
#     for q_a in queryset_combined:
#         survey_data.get(q_a.question).append(q_a.get_content_type_value(q_a.question.get_answer_content_type()))
#
#     for col_num, (question, answer_list) in enumerate(survey_data.items(), 1):
#         row_num = 1
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = question.value
#         cell.font = Font(name='Calibri', bold=True)
#         cell.border = Border(bottom=Side(border_style='medium', color='FF000000'), )
#         column_letter = get_column_letter(col_num)
#         column_dimensions = worksheet.column_dimensions[column_letter]
#         column_dimensions.width = 30
#         for answer in answer_list:
#             row_num += 1
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = answer
#
#     worksheet.freeze_panes = worksheet['A2']
#     workbook.save(response)
#     return response
