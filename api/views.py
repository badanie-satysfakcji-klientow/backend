import datetime
from collections import Counter

from django.db.models import F
from rest_framework import status, serializers
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action

from .serializers import SurveySerializer, SurveyInfoSerializer, ItemSerializer, \
    QuestionSerializer, OptionSerializer, AnswerSerializer, SubmissionSerializer, SectionSerializer, \
    AnswerQuestionCountSerializer, SurveyResultSerializer, SurveyResultInfoSerializer, \
    IntervieweeSerializer, IntervieweeUploadSerializer, SurveyResultFullSerializer, PreconditionSerializer, \
    CreatorSerializer
from .models import Survey, Item, Question, Option, Answer, Submission, Section, Interviewee, Precondition, Creator, \
    SurveySent

from django.http import HttpResponse
import pandas as pd
import csv
from openpyxl import Workbook
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference,
    BarChart
)
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from api.emails import send_my_mass_mail


class SurveyViewSet(ModelViewSet):
    queryset = Survey.objects.prefetch_related('items')
    serializer_class = SurveySerializer
    lookup_url_kwarg = 'survey_id'

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except AttributeError as e:
            return Response({'status': 'error', 'message': e.args})

    @action(detail=False, methods=['GET'], name='Get surveys by creator')
    def retrieve_brief(self, request, *args, **kwargs):  # use prefetch_related
        surveys = Survey.objects.prefetch_related('items', 'items__questions').filter(creator_id=kwargs['creator_id'])
        serializer = SurveyInfoSerializer(surveys, many=True)  # using different serializer for that action
        return Response({'status': 'OK', 'surveys': serializer.data}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['GET'])
    def anonymous_retrieve(self, request, *args, **kwargs):
        survey = Survey.objects.prefetch_related('items')\
            .get(id=SurveySent.objects.get(id=kwargs['survey_hash']).survey_id)
        serializer = self.get_serializer(survey)
        return Response({'status': 'OK', 'survey': serializer.data}, status=status.HTTP_200_OK)


class ItemViewSet(ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    lookup_url_kwarg = 'item_id'

    # TODO: override get_serializer_context

    def create(self, request, *args, **kwargs) -> Response:
        """
        # save survey question by its id
        """
        serializer = self.get_serializer(data=request.data)
        serializer.context['survey_id'] = kwargs.get('survey_id')
        serializer.context['type'] = request.data.get('type')
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        # headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):  # unnecessary
        """
        # update item by its id
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.context['type'] = request.data.get('type')
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data, status=status.HTTP_200_OK)


class AnswersCountViewSet(ModelViewSet):
    serializer_class = AnswerQuestionCountSerializer

    def get_queryset(self):
        submission_queryset = Submission.objects.filter(survey=self.kwargs['survey_id'])
        answers_query = Answer.objects \
            .select_related('question') \
            .filter(submission__in=submission_queryset).values_list('question_id', flat=True)
        return Question.objects.filter(id__in=answers_query)

    def list(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_queryset(), many=True)
        try:
            return Response({'status': 'OK', 'answers_count': serializer.data}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'status': 'Not found', 'message': e.args}, status=status.HTTP_404_NOT_FOUND)
        pass


class SubmissionViewSet(ModelViewSet):
    serializer_class = SubmissionSerializer

    # TODO: override get_serializer_context

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.context['survey_id'] = kwargs.get('survey_id')
        try:
            serializer.is_valid(raise_exception=True)
        except serializers.ValidationError as e:
            return Response({'status': 'error', 'message': e.detail['non_field_errors'][0]},
                            status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        return Response({'status': 'success', 'submission_id': serializer.data.get('id')},
                        status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['POST'])
    def anonymous_create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.context['survey_id'] = SurveySent.objects.get(id=kwargs['survey_hash']).survey_id
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response({'status': 'success', 'submission_id': serializer.data.get('id')},
                        status=status.HTTP_201_CREATED)


class AnswerViewSet(ModelViewSet):
    queryset = Answer.objects.all()
    serializer_class = AnswerSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.context['question_id'] = kwargs.get('question_id')

        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response({'status': 'success', 'answer_id': serializer.data.get('id')},
                        status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = Answer.objects.get(id=kwargs['answer_id'])
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.context['question_id'] = kwargs.get('question_id')
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)


class SectionViewSet(ModelViewSet):
    serializer_class = SectionSerializer
    queryset = Section.objects.all()

    def create(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.context['survey_id'] = kwargs.get('survey_id')
        # validate data
        try:
            serializer.is_valid(raise_exception=True)
        except serializers.ValidationError as e:
            return Response({'status': 'error', 'message': e.detail}, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response({'status': 'created section',
                         'section_id': serializer.data.get('id')},
                        status=status.HTTP_201_CREATED)

    def anonymous_list(self, request, *args, **kwargs):
        survey = SurveySent.objects.get(id=kwargs['survey_hash']).survey

        # TODO: filter so override filter_queryset
        queryset = Section.objects.filter(start_item__in=Item.objects.filter(survey=survey).only('id'))

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class QuestionViewSet(ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    lookup_url_kwarg = 'question_id'

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance_order = instance.order
        instance.delete()
        Question.objects.filter(order__gt=instance_order).update(order=F('order') - 1)
        return Response({'status': 'deleted'}, status=status.HTTP_204_NO_CONTENT)


class OptionViewSet(ModelViewSet):
    queryset = Option.objects.all()
    serializer_class = OptionSerializer

    lookup_url_kwarg = 'option_id'

    def update(self, request, *args, **kwargs):
        """
        # update option by its id
        """
        partial = kwargs.pop('partial', False)
        instance = Option.objects.get(id=kwargs['option_id'])
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            self.perform_update(serializer)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        return Response({'status': 'not updated, wrong parameters'}, status=status.HTTP_400_BAD_REQUEST)


class SurveyResultViewSet(ModelViewSet):
    serializer_class = SurveyResultSerializer
    queryset = Survey.objects.all()
    lookup_url_kwarg = 'survey_id'

    def retrieve(self, request, *args, **kwargs):
        """
        # get survey result by its id
        """
        serializer = self.serializer_class(Question.objects.get(id=self.kwargs['question_id']), many=False)
        return Response({'question_result': serializer.data}, status=status.HTTP_200_OK)

    def list(self, request, *args, **kwargs):
        """
        # get result by survey id
        """
        # get all question results for a survey
        result_info_serializer = SurveyResultInfoSerializer(self.queryset.get(id=self.kwargs['survey_id']), many=False)
        items_query = Item.objects.prefetch_related('questions').filter(survey=self.kwargs['survey_id'])
        result_serializer = self.serializer_class(Question.objects.filter(item_id__in=items_query), many=True)
        return Response({'results_info': result_info_serializer.data,
                         'results': result_serializer.data},
                        status=status.HTTP_200_OK)


class SurveyResultFullViewSet(SurveyResultViewSet):
    serializer_class = SurveyResultFullSerializer


class IntervieweeViewSet(ModelViewSet):
    queryset = Interviewee.objects.all()
    serializer_class = IntervieweeSerializer
    lookup_url_kwarg = 'interviewee_id'


class SendEmailViewSet(ModelViewSet):
    queryset = Interviewee.objects.all()

    @action(detail=False, methods=['POST'])
    def send(self, request, *args, **kwargs):
        """
        send a mail with link to survey
        eg. http://127.0.0.1:4200/survey/survey_uuid
        """
        selected_interviewees = request.query_params.get('selected')
        survey_id = kwargs['survey_id']
        survey_title = Survey.objects.get(id=survey_id).title

        if not selected_interviewees:
            try:
                email_list = request.data['interviewees']
                send_my_mass_mail(survey_id, survey_title, email_list)
            except KeyError as e:
                hint = "Provide correct email list eg. {'interviewees': ['abc@gmail.com', 'kpz@pwr.edu.pl']}"
                return Response(
                    {'status': 'error', 'message': e.args, 'hint': hint}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({'survey_id': survey_id, 'status': 'error', 'message': e.args},
                                status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            return Response({'survey_id': survey_id, 'status': 'sending process started'},
                            status=status.HTTP_200_OK)

        try:
            interviewees = request.data['interviewees']
            selected_interviewees_emails = \
                [self.queryset.get(id=interviewee_id).email for interviewee_id in interviewees]
            send_my_mass_mail(survey_id, survey_title, selected_interviewees_emails)
        except KeyError as e:
            hint = "Provide correct interviewee id list eg.  " \
                   "{'interviewees': ['7d01d6b3-df2a-42fc-ab9e-ffe5f39a9685', '8e813c93-37a7-429f-926c-0ac092b30c79']}"
            return Response(
                {'status': 'error', 'message': e.args, 'hint': hint}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'survey_id': survey_id, 'status': 'error', 'message': e.args},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({'survey_id': survey_id, 'status': 'sending process started'}, status=status.HTTP_200_OK)


class CSVIntervieweesViewSet(ModelViewSet):  # 1. add to db, 2. add to db and send, 3. send without add
    serializer_class = IntervieweeUploadSerializer
    queryset = Interviewee.objects.all()

    @staticmethod
    def get_email_list(already_exists, new_interviewee_list):
        csv_interviewees = already_exists + new_interviewee_list
        return [interviewee.email for interviewee in csv_interviewees]

    @action(detail=False, methods=['POST'])
    def upload_csv(self, request, *args, **kwargs):
        save = request.query_params.get('save')
        survey_id = request.query_params.get('send_survey')

        if not save and not survey_id:
            return Response({'status': 'error', 'message': 'the specified endpoint has no effect',
                             'hint': 'provide available query parameters'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data['file']
        try:
            csv_reader = pd.read_csv(file, sep=';', encoding='utf8')
            new_interviewee_list = []
            already_exists = []
            for idx, row in csv_reader.iterrows():
                new_interviewee = Interviewee(
                    email=row['email'],
                    first_name=row['first_name'],
                    last_name=row['last_name']
                )
                already_exists.append(new_interviewee) if Interviewee.objects.filter(email=row['email']).count() else \
                    new_interviewee_list.append(new_interviewee)
        except pd.errors.EmptyDataError:
            return Response({'status': 'error', 'message': 'selected file is empty or is not .csv',
                             'hint': 'provide CSV with 3 cols with ";" separator eg. email;first_name;last_name'},
                            status=status.HTTP_400_BAD_REQUEST)
        except KeyError:
            return Response({'status': 'error',
                             "message": "selected file doesnt contain 'email', 'first_name' or 'last_name' column",
                             'hint': 'provide CSV with 3 cols: email;first_name;last_name'},
                            status=status.HTTP_400_BAD_REQUEST)

        if survey_id:
            email_list = self.get_email_list(already_exists, new_interviewee_list)
            survey_title = Survey.objects.get(id=survey_id).title
            try:
                send_my_mass_mail(survey_id, survey_title, email_list)
            except Exception as e:
                return Response({'survey_id': survey_id, 'status': 'error during sending email',
                                 'message': e.args}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            if not save:
                return Response({'survey_id': survey_id, 'status': 'sending process started'},
                                status=status.HTTP_200_OK)

            Interviewee.objects.bulk_create(new_interviewee_list)
            return Response(
                {'status': 'respondents saved, sending process started',
                 'newly added': IntervieweeSerializer(new_interviewee_list, many=True).data,
                 'already exists': IntervieweeSerializer(already_exists, many=True).data},  # maybe without existing?
                status=status.HTTP_201_CREATED)

        Interviewee.objects.bulk_create(new_interviewee_list)
        return Response(
            {'status': 'respondents saved',
             'newly added': IntervieweeSerializer(new_interviewee_list, many=True).data,
             'already exists': IntervieweeSerializer(already_exists, many=True).data},  # maybe without existing?
            status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['GET'])
    def download_csv(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="interviewees{datetime.datetime.now()}.csv"'
        response.write(u'\ufeff'.encode('utf8'))

        writer = csv.writer(response)

        header = ';'.join(['email', 'first_name', 'last_name'])
        writer.writerow([header])

        for interviewee in self.queryset:
            row = ';'.join([
                interviewee.email,
                interviewee.first_name,
                interviewee.last_name,
            ])
            writer.writerow([row])

        return response


# for Preconditions
class PreconditionViewSet(ModelViewSet):
    serializer_class = PreconditionSerializer
    lookup_url_kwarg = 'precondition_id'
    queryset = Precondition.objects.all()


class CreatorViewSet(ModelViewSet):
    serializer_class = CreatorSerializer
    lookup_url_kwarg = 'creator_id'
    queryset = Creator.objects.all()
    hint = "Provide correct current_user id eg. " \
           "{'current_user': '8e813c93-37a7-429f-926c-0ac092b30c79'}"

    def check_destroy(self, request, *args, **kwargs):
        if str(request.data.get('current_user')) != str(kwargs.get('creator_id')):
            return Response({'status': 'error', 'message': 'Only creator can delete himself', 'hint': self.hint},
                            status.HTTP_400_BAD_REQUEST)
        return self.destroy(request, *args, **kwargs)

    def check_partial_update(self, request, *args, **kwargs):
        if str(request.data.get('current_user')) != str(kwargs.get('creator_id')):
            return Response({'status': 'error', 'message': 'Only creator can update his data', 'hint': self.hint},
                            status.HTTP_400_BAD_REQUEST)
        return self.partial_update(request, *args, **kwargs)


class QuestionResultRawViewSet(ModelViewSet):
    lookup_url_kwarg = 'question_id'

    def get_queryset(self):
        return Answer.objects.filter(question=Question.objects.get(id=self.kwargs['question_id']))

    def retrieve(self, request, *args, **kwargs):
        question = Question.objects.get(id=self.kwargs['question_id'])
        answer_type = question.get_answer_content_type()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={question}-{date}-results.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'), question=question.value[:15]
        )

        if answer_type == 'option':
            answers = self.get_queryset().prefetch_related('option').values_list('option__content', flat=True)
        elif answer_type == 'content_numeric':
            answers = self.get_queryset().values_list('content_numeric', flat=True)
        elif answer_type == 'content_character':
            answers = self.get_queryset().values_list('content_character', flat=True)
            answers = [' '.join(content_character.lower().strip().split()) for content_character in answers
                       if content_character is not None]
        else:
            return Response({'status': 'error', 'message': 'question without item_type'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        counted = Counter(answers)
        statement = len(counted) > 5
        counted = counted.most_common(5) if statement else counted.most_common()
        max_row = len(counted)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f'"{question.value[:15]}" pie'

        # normal pie
        for row in counted:
            worksheet.append(row)
        pie = PieChart()
        labels = Reference(worksheet, min_col=1, min_row=1, max_row=max_row)
        data = Reference(worksheet, min_col=2, min_row=1, max_row=max_row)
        pie.add_data(data)
        pie.set_categories(labels)
        pie.title = question.value
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showVal = True
        worksheet.add_chart(pie, "D1")

        # projected pie
        ws = workbook.create_sheet(title=f'"{question.value[:15]}" projection')
        for row in counted:
            ws.append(row)
        projected_pie = ProjectedPieChart()
        projected_pie.type = "bar"
        projected_pie.splitType = "pos"  # split by value
        labels = Reference(ws, min_col=1, min_row=1, max_row=max_row)
        data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
        projected_pie.add_data(data)
        projected_pie.dataLabels = DataLabelList()
        projected_pie.dataLabels.showVal = True
        projected_pie.set_categories(labels)
        projected_pie.title = question.value
        ws.add_chart(projected_pie, "D2")

        # bar
        ws2 = workbook.create_sheet(title=f'"{question.value[:15]}" bar')
        for row in counted:
            ws2.append(row)
        bar = BarChart()
        bar.type = "col"
        # bar.style = 10
        bar.title = question.value
        bar.y_axis.title = "Ilość wystąpień"
        bar.x_axis.title = "Udzielone odpowiedzi"
        labels = Reference(ws2, min_col=1, min_row=1, max_row=max_row)
        data = Reference(ws2, min_col=2, min_row=1, max_row=max_row)
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
        bar.add_data(data)
        bar.set_categories(labels)
        bar.shape = 4
        ws2.add_chart(bar, "D3")

        workbook.save(response)
        return response


class SurveyResultRawViewSet(ModelViewSet):
    lookup_url_kwarg = 'survey_id'

    def get_queryset(self):
        items_query = Item.objects.prefetch_related('questions').filter(survey=self.kwargs['survey_id'])
        question_query = Question.objects.filter(item_id__in=items_query)
        return question_query

    def retrieve(self, request, *args, **kwargs):
        survey = Survey.objects.get(id=self.kwargs['survey_id'])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={survey_title}-{date}-results.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'), survey_title=survey.title[:10]
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f'"{survey.title[:15]}" results'

        column_titles = [question.value for question in self.get_queryset()]
        row_num = 1
        for col_num, column_title in enumerate(column_titles, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name='Calibri', bold=True)
            cell.border = Border(bottom=Side(border_style='medium', color='FF000000'),)
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 30

        row_num = 2
        col_num = 1
        for question in self.get_queryset():
            answer_queryset = Answer.objects.filter(question_id=question.id)
            col_data = [a.content_character if a.content_character else a.content_numeric for a in answer_queryset]

            for cell_value in col_data:
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                row_num += 1

            row_num = 2
            col_num += 1

        worksheet.freeze_panes = worksheet['A2']
        workbook.save(response)
        return response
